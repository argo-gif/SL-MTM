import sys
import os

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed yet")
    sys.exit(0)

def main():
    file_path = "uploaded_active_dataset.xlsx"
    print(f"Opening {file_path} with openpyxl (read_only=True)...")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print("Sheets:", wb.sheetnames)
    
    sheet = wb.active
    rows = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        rows.append(row)
        if i >= 5:
            break
            
    print("\nHeader (Row 1):")
    for idx, col in enumerate(rows[0]):
        print(f"Col {idx}: {col}")
        
    print("\nRow 2 Sample Data:")
    for idx, val in enumerate(rows[1]):
        print(f"Col {idx} ({rows[0][idx]}): {val}")

if __name__ == "__main__":
    main()
